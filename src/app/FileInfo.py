# -*- coding: utf-8 -*-

'''
檔案說明：MasterFile、DealerInfo、KAList等檔案初始化建立、合併
Writer：Qian
'''

# masterfile、kalist檔案檢查、維護區塊待寫

# 標準庫
import os, shutil
from datetime import datetime

# 第三方庫
import pandas as pd
from openpyxl import Workbook, load_workbook
from workalendar.asia import Taiwan

# 自定義函數
from SystemConfig import WriteFileJson, WriteDealerJson, SubRecordJson, WrtieWorkDay
from Log import WSysLog
from Config import AppConfig
from __init__ import ConfigJsonFile

Config = AppConfig()
ConfigInfo = ConfigJsonFile()

# 產生對應的 Excel Column 名稱(一次多個)
def get_excel_colmun_name(file_header):
    return [chr(i % 26+ 65) for i in range(len(file_header))]

# 根據表頭搜尋出 Excel 的 Column 名稱(一次一個)
def search_column_name(file_header, col_name):
    for index, value in enumerate(file_header):
        if value == col_name:
            return chr(index % 26 + 65)

# 設定表格樣式
def excel_style(ws, column_width, fixed_columns):
    for i, width in enumerate(column_width):
        ws.column_dimensions[fixed_columns[i]].width = width
    for col in fixed_columns:
        for row in range(1, ws.max_row + 1):
            ws[f"{col}{row}"].alignment = Config.ExcelStyle

# 從使用者資訊中取出BA名單
def get_BA():
    user = Config.UserConfig
    ba_list = []
    for user_id in user:
        if user[user_id]["Group"] == "BD_BA":
            ba_id_part = user[user_id]["Description"].split(" ")
            ba_id = ba_id_part[1]
            ba_name = user[user_id]["Name"]
            ba_dealer_id = user[user_id]["ResponsibleDealerID"]
            ba_list.append({"BA_ID":ba_id, "BA_Name":ba_name, "BA_Dealer_ID" : ba_dealer_id})
    return ba_list

# 依照檔案格式建立空白內容之檔案
def make_format_file(folder_path, file_name, sheet_name, file_header, column_width):
    try:
        file_path = os.path.join(folder_path, file_name)
        wb = Workbook()
        wb.remove(wb.active)
        ws = wb.create_sheet(title = sheet_name)
        fixed_columns = get_excel_colmun_name(file_header)
        for col, data in zip(fixed_columns, file_header):
            ws[f"{col}1"] = data
        excel_style(ws, column_width, fixed_columns)
        wb.save(file_path)
        msg = f"成功於 {folder_path} 目錄中建立空白 {file_name} 檔案。"
        WSysLog("1", "MakeFormatFile", msg)
    except (FileNotFoundError, IOError) as e:
        msg = f"檔案操作錯誤：於 {folder_path} 建立空白 {file_name} 檔案時發生錯誤： {e}。"
        WSysLog("3", "MakeFormatFile", msg)
    except Exception as e:
        msg = f"於 {folder_path} 建立空白 {file_name} 檔案時發生錯誤： {e}。"
        WSysLog("3", "MakeFormatFile", msg)

# 將BA負責的經銷商ID寫入DealerInfo
def wright_dealer_info_in_file(file_path, sheet_name, file_header, ba_dealer_id):
    try:
        wb = load_workbook(file_path)
        ws = wb[sheet_name]
        position_list = ["聯絡人1","聯絡人2","聯絡人3","聯絡人4"]
        content_col = ["Position", "Name", "Mail", "Ex"]
        content_range = 4
        for row in range(2, (len(ba_dealer_id))*content_range+1, content_range):
            index = int((row - 2) / content_range)
            col =  search_column_name(file_header, "Dealer ID")
            ws[f"{col}{row}"] = ba_dealer_id[index]
            
            for j in range(content_range):
                col = search_column_name(file_header, "Position")
                ws[f"{col}{row + j}"] = position_list[j]
                for col_name in content_col:
                    col = search_column_name(file_header, col_name)
                    ws[f"{col}{row + j}"].alignment = Config.ExcelStyle
            
            for col_name in file_header:
                if col_name in content_col:
                    continue
                col = search_column_name(file_header, col_name)
                ws.merge_cells(f"{col}{row}:{col}{row + j}")
                ws[f"{col}{row}"].alignment = Config.ExcelStyle
        wb.save(file_path)
        msg = f"成功寫入經銷商資訊至 {file_path}。"
        WSysLog("1", "WrightDealerInfoInFile", msg)
    except Exception as e:
        msg = f"寫入經銷商資訊發生錯誤： {e}。"
        WSysLog("3", "WrightDealerInfoInFile", msg)

# 檢查BA目錄底下檔案
def CheckBAFolderFiles():
    ba_list = get_BA()
    # 初始建立旗幟
    flag = False
    master_file = Config.MasterFileName
    dealer_info_file = Config.DealerInfoFileName
    kalist_file = Config.KAListFileName
    json_data = Config.FileConfig
    file_time_dic = {}
    for ba in ba_list:
        file_time = {}
        ba_id = ba["BA_ID"]
        ba_name = ba["BA_Name"]
        ba_dealer_id = ba["BA_Dealer_ID"]
        new_ba_id = ba_id[:2] + "_" + ba_id[2:]
        ba_folder_name = new_ba_id + "_" + ba_name
        folder_path = os.path.join(Config.BAFolderPath, ba_folder_name)

        if not os.path.exists(folder_path):
            os.makedirs(folder_path)

        master_file_path = os.path.join(folder_path, master_file)
        dealer_info_file_path = os.path.join(folder_path, dealer_info_file)

        # 目標目錄找不到 MasterFile 時須建立一份空的 MasterFile
        if not os.path.exists(master_file_path):
            flag = True
            msg = f"{ba_folder_name}目錄下缺少{master_file}檔案，系統將建立空白檔案。"
            WSysLog("2", "CheckBAFolderFiles", msg)
            sheet_name = Config.MasterFileSheetName
            file_header = Config.MasterFileHeader
            column_width = Config.MasterFileColumnWidth
            make_format_file(folder_path, master_file, sheet_name, file_header, column_width)
            file_write_time = datetime.fromtimestamp(os.path.getmtime(master_file_path))
            file_write_time = file_write_time.isoformat()
            json_data["FileInfo"][ba_id]["MasterFile"] = file_write_time
            msg = WriteFileJson(json_data)
            WSysLog("1", "CheckBAFolderFiles", msg)
            file_time["MasterFile"] = None
        else:
            file_time["MasterFile"] = datetime.fromtimestamp(os.path.getmtime(master_file_path))

        # 目標目錄找不到 DealerInfo 時須建立一份空的 DealerInfo
        if not os.path.exists(dealer_info_file_path):
            flag = True
            msg = f"{ba_folder_name}目錄下缺少{dealer_info_file}檔案，系統將建立空白檔案。"
            WSysLog("2", "CheckBAFolderFiles", msg)
            sheet_name = Config.DealerInfoFileSheetName
            file_header = Config.DealerInfoFileHeader
            file_header = file_header[1:]
            column_width = Config.DealerInfoFileColumnWidth
            column_width = column_width[1:]
            make_format_file(folder_path, dealer_info_file, sheet_name, file_header, column_width)
            wright_dealer_info_in_file(dealer_info_file_path, sheet_name, file_header, ba_dealer_id)
            file_write_time = datetime.fromtimestamp(os.path.getmtime(dealer_info_file_path))
            file_write_time = file_write_time.isoformat()
            json_data["FileInfo"][ba_id]["DealerInfo"] = file_write_time
            msg = WriteFileJson(json_data)
            WSysLog("1", "CheckBAFolderFiles", msg)
            file_time["DealerInfo"] = None
        else:
            file_time["DealerInfo"] = datetime.fromtimestamp(os.path.getmtime(dealer_info_file_path))

        # 目標目錄找不到 KAList 時須建立一份空的 KAList
        for ka_dealer in Config.KADealerList:
            if ka_dealer in ba_dealer_id:
                ka_list_file_path = os.path.join(folder_path, kalist_file)
                
                # 目標目錄找不到kalist時須建立一份空的kalist
                if not os.path.exists(ka_list_file_path):
                    flag = True
                    msg = f"{ba_folder_name}目錄下缺少{kalist_file}檔案，系統將建立空白檔案。"
                    WSysLog("2", "CheckBAFolderFiles", msg)
                    sheet_name = Config.KAListFileSheetName
                    file_header = Config.KAListFileHeader
                    for i, dealer in enumerate(Config.DealerList):
                        if dealer == ka_dealer:
                            index = i + 1
                            break
                    dealer_name = Config.DealerConfig[f"Dealer{index}"]["DealerName"]
                    sheet_name = ka_dealer + "_" + sheet_name
                    file_header[0] = dealer_name + file_header[0]
                    column_width = Config.KAListFileColumnWidth
                    make_format_file(folder_path, kalist_file, sheet_name, file_header, column_width)
                    file_write_time = datetime.fromtimestamp(os.path.getmtime(ka_list_file_path))
                    file_write_time = file_write_time.isoformat()
                    json_data["FileInfo"][ba_id]["KAList"] = file_write_time
                    msg = WriteFileJson(json_data)
                    file_time["KAList"] = None
                else:
                    file_time["KAList"] = datetime.fromtimestamp(os.path.getmtime(ka_list_file_path))

        file_time_dic[ba_id] = file_time
    
    if not flag:
        msg = "BA目錄底下應有檔案都存在。"
        WSysLog("1", "CheckBAFolderFiles", msg)
        return file_time_dic
    else:
        return file_time_dic
    
# 依據BA人數調整files.json格式
def check_ba():
    flag = True
    ba_list = get_BA()
    json_data = Config.FileConfig

    for ba in ba_list:
        ba_id = ba["BA_ID"]
        if ba_id not in json_data["FileInfo"]:
            flag = False
            json_data["FileInfo"][ba_id] = {"MasterFile":None, "DealerInfo":None, "KAList":None}

    if flag:
        msg = "files.json資料無異動。"
        WSysLog("1", "CheckBA", msg)
        return json_data
    else:
        msg = WriteFileJson(json_data)
        WSysLog("1", "CheckBA", msg)
        return json_data

# 檢查 ba 目錄底下 masterfile 檔案 ##
def check_master_file(check_list):
    print()

# 系統合併 masterfile 檔案
def MergeMasterFile(write_new_list):
    create_flag = False
    ba_list = get_BA()
    ba_folder = Config.BAFolderPath
    folder_path = Config.MasterFolderPath
    file_name = Config.MasterFileName
    file_header = Config.MasterFileHeader
    sheet_name = Config.MasterFileSheetName
    column_width = Config.MasterFileColumnWidth
    general_data_path = os.path.join(folder_path, file_name)

    # 無 Master File 總表
    if not os.path.exists(general_data_path):
        msg = f"{folder_path} 目錄底下不存在 {file_name} 總表檔案，系統將重新產生。"
        WSysLog("2", "MergeKAList", msg)
        wb = Workbook()
        wb.remove(wb.active)
        wb.create_sheet(title = sheet_name)
        create_flag = True
        wb.save(general_data_path)

    # 有 Master File 總表，無工作表
    else:
        wb = load_workbook(general_data_path)
        # MasterFile 總表中無 MasterFile 工作表
        if sheet_name not in wb.sheetnames:
            msg = f"{file_name} 總表檔案中 {sheet_name} 工作表不存在，系統將重新產生。"
            WSysLog("2", "MergeKAList", msg)
            wb.create_sheet(title = sheet_name)
            create_flag = True
            wb.save(general_data_path)

    wb = load_workbook(general_data_path)
    ws = wb[sheet_name]

    # 在全新工作表中寫入合併後的資料
    if create_flag:
        check_master_file("all")
        dfs = []
        msg = f"系統在 {file_name} 總表檔案中，產生 {sheet_name} 工作表內容。"
        WSysLog("1", "MergeKAList", msg)

        for ba in ba_list:
            ba_folder_name = ba["BA_ID"][:2] + "_" + ba["BA_ID"][2:] + "_" + ba["BA_Name"]
            file_path = os.path.join(ba_folder, ba_folder_name, file_name)

            # 檢查ba目錄底下各自的 MasterFile 檔案
            if os.path.exists(file_path):
                file = pd.ExcelFile(file_path)

                # 檢查檔案中是否存在 MasterFile 工作表 
                if sheet_name in file.sheet_names:
                    df = pd.read_excel(file_path, sheet_name = sheet_name, names = file_header)
                    if df.empty:
                        continue
                    dfs.append(df)

                else:
                    msg = "MasterFile 工作表不存在於檔案中。"
                    WSysLog("3", "MegerMasterFile", msg)
                    raise NameError(msg)

            else:
                msg = f"該 ba {ba['BA_Name']} 目錄底下無 MasterFile 檔案。"
                WSysLog("3", "MegerMasterFile", msg)
                raise FileNotFoundError(msg)

        # 內容合併
        master_data = pd.concat(dfs, ignore_index = True)

    # 更新現有工作表內容
    else:
        check_master_file(write_new_list)
        master_data = pd.read_excel(general_data_path, sheet_name = sheet_name, dtype = str)
        file_header = master_data.columns.values
        dealer_ids = sorted(list(set(master_data[file_header[0]].values)))
        part_data = {dealer_id: master_data[master_data[file_header[0]] ==\
                    dealer_id].reset_index(drop=True) for dealer_id in dealer_ids}
        
        for ba_id in write_new_list:
            print(ba_id)
            for ba in ba_list:
                dealer_list_in_file = None
                if ba_id == ba["BA_ID"]:
                    ba_folder_name = ba["BA_ID"][:2] + "_" + ba["BA_ID"][2:] + "_" + ba["BA_Name"]
                    master_file_path = os.path.join(ba_folder, ba_folder_name, file_name)
                    file_data = pd.read_excel(master_file_path, sheet_name = sheet_name, dtype = str)
                    dealer_list_in_file = sorted(list(set(file_data[file_header[0]].values)))
                    part_data_in_ba_file = {dealer_id: file_data[file_data[file_header[0]] ==\
                                            dealer_id].reset_index(drop = True) for \
                                            dealer_id in dealer_list_in_file}
                if dealer_list_in_file:
                    for dealer_id in dealer_list_in_file:
                        # 若經銷商id值不在總表中，直接新增
                        if dealer_id not in dealer_ids:
                            part_data[dealer_id] = part_data_in_ba_file[dealer_id]
                        # 若經銷商id值在總表中，逐筆比對貨號之區間值
                        else:
                            general_data = part_data[dealer_id]
                            ba_data = part_data_in_ba_file[dealer_id]
                            for _, row in ba_data.iterrows():
                                
                                # 比對產品ID及起迄區間
                                part_general = general_data[(general_data[file_header[1]] == row.iloc[1]) &\
                                                            (general_data[file_header[-2]] == row.iloc[-2]) &\
                                                            (general_data[file_header[-1]] == row.iloc[-1])]

                                # 區間值符合一項
                                if len(part_general) == 1:
                                    data_key = file_header[2:-2]
                                    for key in data_key:
                                        if str(part_general[key].values[0]) != str(row[key]):
                                            index = part_general.index.values[0]
                                            part_data[dealer_id].loc[index, key] = row[key]
                                            msg = f"更新 MasterFile 總表 {dealer_id} {row.iloc[1]} {key} 的值為： {row[key]}。"
                                            WSysLog("1", "MergeKalist", msg)

                                # 區間值未符合
                                elif part_general.empty:
                                    msg = f"{dealer_id} 之產品ID： {row.iloc[1]} 為搜尋到對應的起迄值。"
                                    WSysLog("1", "MergeMasterFile", msg)
                                    new_row = row.to_dict()
                                    max_row = len(part_data[dealer_id])
                                    part_data[dealer_id][max_row] = new_row
                                    msg = f"新增資料： {new_row} 至 {max_row + 2}row。"
                                    WSysLog("1", "MergeMasterFile", msg)

                                # 符合區間值過多
                                else:
                                    msg = f"{dealer_id} 之產品ID： {row.iloc[1]} 在同一起迄區間擁有多個值。"
                                    WSysLog("3", "MergeMasterFile", msg)

        # 合併表格
        master_data = pd.concat(part_data, ignore_index = True)

    # 表頭填寫
    file_header = master_data.columns.values
    fixed_columns = get_excel_colmun_name(file_header)
    for col, data in zip(fixed_columns, file_header):
        ws[f"{col}1"] = data
    excel_style(ws, column_width, fixed_columns)

    # 內容填寫
    for row in range(len(master_data)):
        for col_name in file_header:
            col = search_column_name(file_header, col_name)
            ws[f"{col}{row + 2}"] = master_data[col_name][row]
            ws[f"{col}{row + 2}"].alignment = Config.ExcelStyle
    wb.save(general_data_path)
    msg = "MasterFile 檔案中 MasterFile 工作表維護完成。"
    WSysLog("1", "MergeMasterFile", msg)

# 將各BA目錄下小表中的經銷商表頭拷貝至總表
def MergeDealerInfoWorkSheet(write_new_list):
    ba_folder = Config.BAFolderPath
    folder_path = Config.DealerInfoPath
    file_name = Config.DealerInfoFileName
    sheet_name = Config.DealerInfoFileSheetName
    ba_list = get_BA()
    for ba_id in write_new_list:
        for ba in ba_list:
            if ba["BA_ID"] == ba_id:
                ba_folder_name = ba["BA_ID"][:2] + "_" + ba["BA_ID"][2:] + "_" + ba["BA_Name"]
                dealers = ba["BA_Dealer_ID"]
                break

        # 各 BA 目錄底下小表
        file_path = os.path.join(ba_folder, ba_folder_name, file_name)
        file_data = load_workbook(file_path)
        file_sheets = file_data.sheetnames
        file_sheets.remove(sheet_name)

        # 總表
        target_path = os.path.join(folder_path, file_name)
        target_data = load_workbook(target_path)
        target_sheets = target_data.sheetnames

        for dealer_id in dealers:
            sSheetName = f"{dealer_id}_Sale"
            iSheetName = f"{dealer_id}_Inventory"
            if (sSheetName not in file_sheets) & (sSheetName in target_sheets):
                target_data.remove(target_data[sSheetName])
            msg = f"{sSheetName} 僅存在於總表中，系統將刪除。"
            WSysLog("1", "MergeDealerInfoWordSheet", msg)
            if (iSheetName not in file_sheets) & (iSheetName in target_sheets):
                target_data.remove(target_data[iSheetName])
            msg = f"{iSheetName} 僅存在於總表中，系統將刪除。"
            WSysLog("1", "MergeDealerInfoWordSheet", msg)

        for sheet in file_sheets:
            file_ws = file_data[sheet]
            if sheet in target_data.sheetnames:
                target_data.remove(target_data[sheet])
            target_ws = target_data.create_sheet(title = sheet)

            for row in file_ws.iter_rows():
                for cell in row:
                    target_ws[cell.coordinate].value = cell.value
    
        target_data.save(target_path)
        msg = f"更新 DealerInfo 總表中的 {file_sheets} 工作表。"
        WSysLog("1", "MergeDealerInfoWordSheet", msg)

# 系統合併 DealerInfo 檔案
def MergeDealerInfo(write_new_list):
    ba_list = get_BA()
    dealer_list = Config.DealerList
    ba_folder = Config.BAFolderPath
    folder_path = Config.DealerInfoPath
    file_name = Config.DealerInfoFileName
    sheet_name = Config.DealerInfoFileSheetName
    file_header = Config.DealerInfoFileHeader
    column_width = Config.DealerInfoFileColumnWidth
    general_data_path = os.path.join(folder_path, file_name)
    content_col = ["Position", "Name", "Mail", "Ex"]

    # 總表存放目錄底下沒有總表檔案時，直接合併全部產生一份檔案
    if not os.path.exists(general_data_path):
        write_new_list.clear()
        msg = f"{folder_path} 目錄底下無 {file_name} 總表檔案，系統將重新產生。"
        WSysLog("2", "MergeDealerInfo", msg)
        wb = Workbook()
        wb.remove(wb.active)
        ws = wb.create_sheet(title = sheet_name)
        dataframes, dealers, dealer_part, dealer_part_dct = [], [], [], {}

        for ba in ba_list:
            write_new_list.append(ba["BA_ID"])
            ba_folder_name = ba["BA_ID"][:2] + "_" + ba["BA_ID"][2:] + "_" + ba["BA_Name"]
            file_path = os.path.join(ba_folder, ba_folder_name, file_name)
            dealer_data = pd.read_excel(file_path, sheet_name = sheet_name, dtype = str)
            dealer_id_in_data = dealer_data[dealer_data["Dealer ID"].notna()]
            dealer_id_in_data = dealer_id_in_data["Dealer ID"].values
            part = [dealer_data.iloc[i:i+4].reset_index(drop=True)\
                     for i in range(0, len(dealer_data), 4)]
            dealer_part.extend(part)
            dealers.extend(ba["BA_Dealer_ID"])

            # 不該存在的Dealer，存在檔案中
            for dealer in dealer_id_in_data:
                if dealer not in ba["BA_Dealer_ID"]:
                    msg = f" {dealer} 不存在於BA負責範圍，請系統管理員修正。"
                    WSysLog("3", "MergeDealerInfo", msg)

                    df_index = dealer_data[dealer_data["Dealer ID"] == dealer].index
                    for j in range(4):
                        dealer_data = dealer_data.drop(index = [int(df_index.values[0]) + j])

            for dealer in ba["BA_Dealer_ID"]:
                # 應該存在的DealerID不再檔案中
                if dealer not in dealer_data["Dealer ID"].values:
                    msg = f"{dealer} 資訊不存在於 {ba['BA_ID']} 的 {file_name} 檔案中，請系統管理員修正 DealerInfo 檔案。"
                    WSysLog("2", "MergeDealerInfo", msg)

                if dealer not in dealer_list:
                    dealer_list.append(dealer)
                    result = WriteDealerJson("DealerList", dealer_list)
                    if result:
                        msg = f"將 {dealer} 加入 DealerList 紀錄中。"
                        WSysLog("1", "WriteDealerJson", msg)
                        msg = SubRecordJson("Start", None)
                        WSysLog("1", "SubRecordJson", msg)

        dealer_part_dct = {part["Dealer ID"][0]:part for part in dealer_part}

        for dealer in dealer_part_dct:
            dealer_part_dct[dealer].insert(0, "ID", None)
            dealer_part_dct[dealer].loc[0, "ID"] = \
                f"Dealer{(dealer_list.index(dealer_part_dct[dealer]['Dealer ID'][0]) + 1)}"
            dataframes.append(dealer_part_dct[dealer])

        conbined_df = pd.concat(dataframes, ignore_index = True)
        fixed_columns = get_excel_colmun_name(file_header)
        for col, data in zip(fixed_columns, file_header):
            ws[f"{col}1"] = data
        excel_style(ws, column_width, fixed_columns)

        # 將資料寫入工作表
        for row in range(len(conbined_df)):
            for col_name in conbined_df.columns.values:
                col = search_column_name(file_header, col_name)
                ws[f"{col}{row + 2}"] = conbined_df[col_name][row]
                ws[f"{col}{row + 2}"].alignment = Config.ExcelStyle

        for row in range(2, len(conbined_df) + 1, 4):
            for col_name in conbined_df.columns.values:
                if col_name not in content_col:
                    col = search_column_name(file_header, col_name)
                    ws.merge_cells(f"{col}{row}:{col}{row + 3}")
        wb.save(general_data_path)

    # 總表存在，做單欄位修正
    else:
        wb = load_workbook(general_data_path)
        ws = wb[sheet_name]
        dealers, dealer_part, dealer_part_dct = [], [], {}
        new_flag = False

        for ba_id in write_new_list:
            for ba in ba_list:
                dealers.extend(ba["BA_Dealer_ID"])
                if ba["BA_ID"] == ba_id:
                    ba_dealer_id = ba["BA_Dealer_ID"]
                    ba_folder_name = ba["BA_ID"][:2] + "_" + ba["BA_ID"][2:] + "_" + ba["BA_Name"]
                    break
            file_path = os.path.join(Config.BAFolderPath, ba_folder_name, file_name)
            dealer_data = pd.read_excel(file_path, sheet_name = sheet_name, dtype = str)
            dealer_id_in_data = dealer_data[dealer_data["Dealer ID"].notna()]

            # BA目錄底下的DealerInfo之dealerid值不存在於user.json中，系統自動略過
            for dealer in dealer_id_in_data["Dealer ID"]:
                if dealer not in ba_dealer_id:
                    msg = f" {dealer} 不存在於BA負責範圍，請系統管理員修正。"
                    WSysLog("3", "MergeDealerInfo", msg)
                    df_index = dealer_data[dealer_data["Dealer ID"] == dealer].index
                    for j in range(4):
                        dealer_data = dealer_data.drop(index = [int(df_index.values[0]) + j])

            # 經銷商ID存在於使用者json中，但不存在於Dealer.json，將資訊加入至Dealer.json
            for dealer in ba_dealer_id:
                if dealer not in dealer_data["Dealer ID"].values:
                    msg = f"{dealer} 經銷商資訊不存在於 {ba['BA_ID']} 的 {file_name} 檔案中，請系統管理員修正 DealerInfo 檔案。"
                    WSysLog("2", "MergeDealerInfo", msg)

                if dealer not in dealer_list:
                    dealer_list.append(dealer)
                    result = WriteDealerJson("DealerList", dealer_list)

                    if result:
                        msg = f"將 {dealer} 加入 DealerList 紀錄中。"
                        WSysLog("1", "WriteDealerJson", msg)
                        msg = SubRecordJson("Start", None)
                        WSysLog("1", "SubRecordJson", msg)

            dealer_data = dealer_data.apply(lambda col:\
                col.map(lambda x: None if pd.isna(x) else x))
            part = [dealer_data.iloc[i:i+4].reset_index(drop=True)\
                for i in range(0, len(dealer_data), 4)]
            dealer_part.extend(part)

        # df資料轉變為字典型態
        dealer_part_dct = {part["Dealer ID"][0]:part for part in dealer_part}
        for dealer in dealer_part_dct:
            # 檢查異動檔案中的經銷商ID是否符合 BA 負責的
            if dealer not in dealers:
                msg = f"檔案中的 {dealer} 資訊未存在於 BA 負責的經銷商 ID 中，請系統管理員維護 BA 負責之經銷商。"
                WSysLog("2", "MergeDealerInfo", msg)

            dealer_part_dct[dealer].insert(0, "ID", None)
            dealer_part_dct[dealer].loc[0, "ID"] =\
                f"Dealer{(dealer_list.index(dealer_part_dct[dealer]['Dealer ID'][0]) + 1)}"
            index = dealer_list.index(dealer)
            row = (index * 4) + 2
            for col_name in dealer_part_dct[dealer].columns.values:
                col = search_column_name(file_header, col_name)
                for i in range(len(dealer_part_dct[dealer])):
                    if ws[f"{col}{row + i}"].value != dealer_part_dct[dealer][col_name][i]:
                        new_flag = True
                        msg = f"更新總表 {col}{row + i} 資訊為： {dealer_part_dct[dealer][col_name][i]}。"
                        WSysLog("1", "MergeDealerInfo", msg)
                        ws[f"{col}{row + i}"] = dealer_part_dct[dealer][col_name][i]
                        ws[f"{col}{row + i}"].alignment = Config.ExcelStyle
                if col_name not in content_col:
                    ws.merge_cells(f"{col}{row}:{col}{row + 3}")

        wb.save(general_data_path)
        if not new_flag:
            msg = "Dealer 資料無異動。"
            WSysLog("1", "MergeDealerInfo", msg)

    MergeDealerInfoWorkSheet(write_new_list)

# 依據總表更新Dealer.json檔案
def RenewDealerJson():
    dealer_config = Config.DealerConfig
    file_path = os.path.join(Config.DealerInfoPath, Config.DealerInfoFileName)
    sheet_name = Config.DealerInfoFileSheetName

    # header 拆開處理
    dealer_info_header = Config.DealerInfoFileHeader[1:8]
    dealer_connect_header = Config.DealerInfoFileHeader[8:12]
    dealer_ka_header = Config.DealerInfoFileHeader[12]
    dealer_OUP_header = Config.DealerInfoFileHeader[15]
    dealer_sale_file_header = Config.DealerInfoFileHeader[13:16]
    dealer_inventory_file_header = Config.DealerInfoFileHeader[16:18]

    dealer_data = pd.read_excel(file_path, sheet_name = sheet_name, dtype = str)
    dealer_data = dealer_data.apply(lambda col:\
        col.map(lambda x: None if pd.isna(x) else x))
    dealer_data_part = [dealer_data.iloc[i:i+4].reset_index(drop=True)\
        for i in range(0, len(dealer_data), 4)]
    dealer_part_dct = {part["ID"][0]:part for part in dealer_data_part}

    ka_list = []
    for dealer in dealer_part_dct:
        dealer_json = {}
        dealer_id = dealer_part_dct[dealer][dealer_info_header[0]][0]
        sale_file_sheet_name = f"{dealer_id}_Sale"
        inventory_file_sheet_name = f"{dealer_id}_Inventory"
        sheets = pd.ExcelFile(file_path).sheet_names
        dealer_sale_format = pd.read_excel(file_path, \
                                sheet_name = sale_file_sheet_name, dtype = str)\
                                if sale_file_sheet_name in sheets else pd.DataFrame()

        dealer_inventory_format = pd.read_excel(file_path,\
                                    sheet_name = inventory_file_sheet_name, dtype = str)\
                                    if inventory_file_sheet_name in sheets else pd.DataFrame()
        # 取出總表中的 ka 值
        ka_data = dealer_part_dct[dealer].loc[0, dealer_ka_header]
        if ka_data == "T":
            dealer_part_dct[dealer].loc[0, dealer_ka_header] = True
            ka_list.append(dealer_id)
        else:
            dealer_part_dct[dealer].loc[0, dealer_ka_header] = False

        OUP_data = dealer_part_dct[dealer].loc[0, dealer_OUP_header]
        dealer_part_dct[dealer].loc[0, dealer_OUP_header] = True if OUP_data == "T" else False

        # 取出總表中的經銷商基本資訊
        dealer_info = {}
        for col_name in dealer_info_header:
            dealer_info[col_name.replace(" ","")] = dealer_part_dct[dealer].loc[0, col_name]

        # 取出總表中的經銷商聯絡資訊
        dealer_connect = {}
        for row in range(4):
            for col_name in dealer_connect_header:
                dealer_connect[f"Contact{row + 1}{col_name}"] =\
                    dealer_part_dct[dealer].loc[row, col_name]

        # 取出總表中經銷商銷售檔案資訊
        dealer_sale_file = {}
        for col_name in dealer_sale_file_header:
            dealer_sale_file[col_name.replace(" ", "").replace("SaleFile", "")] =\
                dealer_part_dct[dealer].loc[0, col_name]
        dealer_sale_file["FileHeader"] = dealer_sale_format.columns.values.tolist()

        # 取出總表中經銷商庫存檔案資訊
        dealer_inventory_file = {}
        for col_name in dealer_inventory_file_header:
            dealer_inventory_file[col_name.replace(" ", "").replace("InventoryFile", "")] =\
                dealer_part_dct[dealer].loc[0, col_name]
        dealer_inventory_file["FileHeader"] = dealer_inventory_format.columns.values.tolist()
        
        dealer_file = {"SaleFile":dealer_sale_file, "InventoryFile":dealer_inventory_file}
        dealer_json = {**dealer_info, **dealer_connect, **dealer_file}

        if dealer in dealer_config:
            dealer_config[dealer] = dealer_json
            result = WriteDealerJson("DealerInfo", dealer_config)
            if result:
                msg = f"更新 dealer.json 中的 {dealer} 資訊。"
                WSysLog("1", "RenewDealerJson", msg)
        if dealer not in dealer_config:
            dealer_config[dealer] = dealer_json
            result = WriteDealerJson("DealerInfo", dealer_config)
            if result:
                msg = f"在 dealer.json 中新增 {dealer} 的資訊。"
                WSysLog("1", "RenewDealerJson", msg)

    # 更新 dealer.json 中的 KAList
    original_ka_list = dealer_config["KADealerList"]
    if original_ka_list != ka_list:
        dealer_config["KADealerList"] = ka_list
        result = WriteDealerJson("DealerInfo", dealer_config)
        if result:
            msg = f"更新 dealer.json 中的 KAList：{ka_list}。"
            WSysLog("1", "RenewDealerJson", msg)

# 檢查 ba 目錄底下的 ka 檔案 ##
def check_ka_file(check_list):
    print()

# 系統合併 KAList 檔案
def MergeKalist(write_new_list):
    create_flag = False
    ba_list = get_BA()
    ba_folder = Config.BAFolderPath
    folder_path = Config.MasterFolderPath
    ka_file_name = Config.KAListFileName
    column_width = Config.KAListFileColumnWidth
    file_name = Config.MasterFileName
    sheet_name = Config.KAListFileSheetName
    general_data_path = os.path.join(folder_path, file_name)
    
    # 無 Master File 總表
    if not os.path.exists(general_data_path):
        msg = f"{folder_path} 目錄底下無 {file_name} 總表檔案，系統將重新產生。"
        WSysLog("2", "MergeKAList", msg)
        wb = Workbook()
        wb.remove(wb.active)
        wb.create_sheet(title = sheet_name)
        create_flag = True
        wb.save(general_data_path)

    # 有 Master File 總表，無工作表
    else:
        wb = load_workbook(general_data_path)
        # MasterFile總表中無kalist工作表
        if sheet_name not in wb.sheetnames:
            msg = f"{file_name} 總表檔案中 {sheet_name} 不存在，系統將重新產生。"
            WSysLog("2", "MergeKAList", msg)
            wb.create_sheet(title = sheet_name)
            create_flag = True
            wb.save(general_data_path)

    wb = load_workbook(general_data_path)
    ws = wb[sheet_name]

    # 在全新工作表中寫入合併後的資料
    if create_flag:
        check_ka_file("all")
        dfs = []
        msg = f"系統在 {file_name} 總表檔案中，產生 {sheet_name} 工作表內容。"
        WSysLog("1", "MergeKAList", msg)

        for ba in ba_list:
            ba_folder_name = ba["BA_ID"][:2] + "_" + ba["BA_ID"][2:] + "_" + ba["BA_Name"]
            ka_file_path = os.path.join(ba_folder, ba_folder_name, ka_file_name)

            if os.path.exists(ka_file_path):
                file = pd.ExcelFile(ka_file_path)
                sheets = file.sheet_names

                for sheet in sheets:
                    part = sheet.split("_")
                    dealer_id = part[0]

                    df = pd.read_excel(ka_file_path, sheet_name = sheet, names = Config.KAListFileHeader)
                    df.insert(loc = 0, column = "經銷商ID", value = dealer_id)
                    df[Config.KAListFileHeader[1]] = pd.to_datetime(df[Config.KAListFileHeader[1]])
                    df[Config.KAListFileHeader[2]] = pd.to_datetime(df[Config.KAListFileHeader[2]])
                    df[Config.KAListFileHeader[1]] = df[Config.KAListFileHeader[1]].dt.date
                    df[Config.KAListFileHeader[2]] = df[Config.KAListFileHeader[2]].dt.date
                    dfs.append(df)

        # 內容合併
        ka_data = pd.concat(dfs, ignore_index = True)

    # MasterFile總表中有kalist工作表
    else:
        check_ka_file(write_new_list)
        ka_data = pd.read_excel(general_data_path, sheet_name = sheet_name, dtype = str)
        ka_data_header = ka_data.columns.values
        
        for ba_id in write_new_list:
            for ba in ba_list:
                if ba["BA_ID"] == ba_id:
                    ba_folder_name = ba["BA_ID"][:2] + "_" + ba["BA_ID"][2:] + "_" + ba["BA_Name"]
                    ka_file_path = os.path.join(ba_folder, ba_folder_name, ka_file_name)
                    sheets = pd.ExcelFile(ka_file_path).sheet_names

                    # 執行每個工作表
                    for sheet in sheets:
                        part = sheet.split("_")
                        dealer_id = part[0]
                        ka_data_part = ka_data[ka_data[ka_data_header[0]] == dealer_id]
                        data = pd.read_excel(ka_file_path, sheet_name = sheet, names = Config.KAListFileHeader, dtype = str)
                        data_header = data.columns.values
                        
                        # 歷遍客戶ID
                        for buyer_id in data[data_header[0]]:
                            # 從總表中篩選出對應的客戶id
                            ka_buyer_data = ka_data_part[ka_data_part[ka_data_header[1]] == buyer_id]
                            # 從子表中篩選出對應的客戶id
                            buyer_data = data[data[data_header[0]] == buyer_id]
                            buyer_data = buyer_data.reset_index(drop=True)

                            if not ka_buyer_data.empty:
                                for row in range(len(buyer_data)):
                                    # 篩選出符合起迄區間的資料
                                    map_data = ka_buyer_data[(ka_buyer_data[ka_data_header[2]] == buyer_data[data_header[1]][row])\
                                                            & (ka_buyer_data[ka_data_header[3]] == buyer_data[data_header[2]][row])]
                                    
                                    if len(map_data) == 1:
                                        if map_data[ka_data_header[4]].values[0] != buyer_data[data_header[3]][row]:
                                            value = buyer_data[data_header[3]][row]
                                            ka_row = map_data.index.values[0]
                                            ka_data.loc[ka_row, ka_data_header[4]] = value
                                            msg = f"修正總表 {ka_row + 2} 的值為： {value}。"
                                            WSysLog("1", "MergeKalist", msg)

                                    elif map_data.empty:
                                        msg = "經銷商之客戶ID未搜尋到對應的起迄區間。"
                                        WSysLog("1", "MergeKalist", msg)
                                        new_row = buyer_data.iloc[row].to_dict()
                                        max_row = len(ka_data)
                                        ka_data.loc[max_row] = new_row
                                        msg = f"新增資料至 {max_row + 2}row。"
                                        WSysLog("1", "MergeKalist", msg)

                                    else:
                                        msg = "經銷商之客戶ID在同一起迄區間擁有多個值。"
                                        WSysLog("3", "MergeKalist", msg)

                            else:
                                for row in range(len(buyer_data)):
                                    new_row = buyer_data.iloc[row].to_dict()
                                    max_row = len(ka_data)
                                    ka_data.loc[max_row] = new_row
                                    msg = f"新增 {buyer_id} 資料至 {max_row + 2}row。"
                                    WSysLog("1", "MergeKalist", msg)

    # 表頭填寫
    file_header = ka_data.columns.values
    fixed_columns = get_excel_colmun_name(file_header)
    for col, data in zip(fixed_columns, file_header):
        ws[f"{col}1"] = data
    column_width = [15] + column_width
    excel_style(ws, column_width, fixed_columns)

    # 內容填寫
    for row in range(len(ka_data)):
        for col_name in ka_data.columns.values:
            col = search_column_name(file_header, col_name)
            ws[f"{col}{row + 2}"] = ka_data[col_name][row]
            ws[f"{col}{row + 2}"].alignment = Config.ExcelStyle
    wb.save(general_data_path)
    msg = "MasterFile 檔案中 KAList 工作表維護完成。"
    WSysLog("1", "MegerKAList", msg)

# 檢查檔案資訊主程式
def CheckFileInfo():
    master_new, kalist_new, dealerinfo_new = [], [], []
    os_file_time = CheckBAFolderFiles()
    json_data = check_ba()
    record_file_time = json_data["FileInfo"]
    master_file_path = os.path.join(Config.MasterFolderPath, Config.MasterFileName)
    dealer_file_path = os.path.join(Config.DealerInfoPath, Config.DealerInfoFileName)
    if os.path.exists(master_file_path):
        master_data = load_workbook(master_file_path)
        master_sheets = master_data.sheetnames
    else:
        master_sheets = []

    # 取得作業系統檔案更新時與紀錄中比對，判定是否需要更新
    for BA_id, files_time in os_file_time.items():
        for file_name, file_time in files_time.items():
            if file_time:
                last_file_time = record_file_time[BA_id][file_name]
                if last_file_time:
                    last_file_time = datetime.fromisoformat(last_file_time)
                    if last_file_time != file_time:
                        if file_time > last_file_time:
                            msg = f"{BA_id} 的 {file_name} 檔案內容更新，系統需更新 {file_name} 總表。"
                            WSysLog("1", "CheckFileInfo", msg)
                            file_time = file_time.isoformat()
                            record_file_time[BA_id][file_name] = file_time
                            json_data["FileInfo"] = record_file_time
                            msg = WriteFileJson(json_data)
                            WSysLog("1", "CheckFileInfo", msg)
                            if file_name == "MasterFile":
                                master_new.append(BA_id)
                            elif file_name == "DealerInfo":
                                dealerinfo_new.append(BA_id)
                            elif file_name == "KAList":
                                kalist_new.append(BA_id)

                        else:
                            msg = f"{BA_id} 的 {file_name}檔案異動時間異常，小於系統紀錄時間。"
                            WSysLog("3", "CheckFileInfo", msg)

                    else:
                        msg = f"系統偵測 {BA_id} 的 {file_name} 檔案無更新。"
                        WSysLog("1", "CheckFileInfo", msg)
                else:
                    # 寫入初始化 files.json檔案
                    file_time = file_time.isoformat()
                    record_file_time[BA_id][file_name] = file_time
                    json_data["FileInfo"] = record_file_time
                    msg = WriteFileJson(json_data)
                    WSysLog("1", "CheckFileInfo", msg)

                    if file_name == "MasterFile":
                        master_new.append(BA_id)
                    elif file_name == "DealerInfo":
                        dealerinfo_new.append(BA_id)
                    elif file_name == "KAList":
                        kalist_new.append(BA_id)

            else:
                msg = f"系統初始化建立 {file_name} 檔案於 {BA_id} 中，請填寫完畢後，在執行一次系統。"
                WSysLog("1", "CheckFileInfo", msg)

    if master_new or (not os.path.exists(master_file_path)):
        MergeMasterFile(master_new)
    if dealerinfo_new or (not os.path.exists(dealer_file_path)):
        MergeDealerInfo(dealerinfo_new)
        RenewDealerJson()
    if kalist_new or (Config.KAListFileSheetName not in master_sheets):
        MergeKalist(kalist_new)

# 檢查雲端的Config檔案，將其拷貝至系統運行目錄
def ConfigFile():
    # 取得雲端的config檔案
    could_config_path = Config.ConfigFolderPath
    could_configs = [file for file in os.listdir(could_config_path)\
        if os.path.isfile(os.path.join(could_config_path, file))]
    
    # 取得本地的config檔案
    local_config_path = ConfigInfo.CONFIG_FOLDER
    local_configs = [file for file in os.listdir(local_config_path)\
        if os.path.isfile(os.path.join(local_config_path, file))]
    
    diff = set(could_configs) == set(local_configs)
    if diff:
        for file_name in local_configs:
            could_file_path = os.path.join(could_config_path, file_name)
            local_file_path = os.path.join(local_config_path, file_name)
            could_file_time = datetime.fromtimestamp(os.path.getmtime(could_file_path))
            local_file_time = datetime.fromtimestamp(os.path.getmtime(local_file_path))
            if could_file_time == local_file_time:
                msg = f"系統確認 {file_name} 檔案雲端及本地版本一致。"
                WSysLog("1", "ConfigFile", msg)
            elif could_file_time > local_file_time:
                msg = f"雲端 {file_name} 檔案需同步至本地。"
                WSysLog("1", "ConfigFile", msg)
                try:
                    shutil.copy2(could_file_path, local_file_path)
                    msg = f"系統從雲端目錄 {could_config_path} 拷貝 {file_name} 至本地目錄 {local_config_path} 。"
                    WSysLog("1", "ConfigFile_SyncConfigs", msg)
                except (IOError, shutil.Error) as e:
                    msg = f"拷貝 {file_name} 發生未知錯誤，錯誤原因： {e} 。"
                    WSysLog("3", "ConfigFile", msg)
            else:
                msg = f"本地 {file_name} 檔案版本大於雲端，系統將直接使用。"
                WSysLog("1", "ConfigFile", msg)
        return True
    else:
        msg = "雲地兩端 Config 目錄中檔案不相等，請系統管理員進行維護。"
        WSysLog("3", "ConfigFile", msg)
        diff = list(set(could_configs) - set(local_configs))
        if diff:
            msg = f"比對雲地兩端Config目錄，雲端 Config 目錄中有 {'、'.join(diff)} 檔案不存在於本地中。"
            WSysLog("3", "ConfigFile", msg)
        diff = list(set(local_configs) - set(could_configs))
        if diff:
            msg = f"比對雲地兩端Config目錄，本地 Config 目錄中有 {'、'.join(diff)} 檔案不存在於雲端中。"
            WSysLog("3", "ConfigFile", msg)
        return False

# 將系統運行的config檔案上傳至雲端
def ConfigFileToCould():
    # 取得雲端的config檔案
    could_config_path = Config.ConfigFolderPath
    
    # 取得本地的config檔案
    local_config_path = ConfigInfo.CONFIG_FOLDER
    local_configs = [file for file in os.listdir(local_config_path)\
        if os.path.isfile(os.path.join(local_config_path, file))]
    try:
        for config_name in local_configs:
            local_config = os.path.join(local_config_path, config_name)
            could_config = os.path.join(could_config_path, config_name)
            if not os.path.exists(could_config) or os.stat(local_config).st_mtime > os.stat(could_config).st_mtime:
                shutil.copy2(local_config, could_config)
                msg = f"拷貝Config檔案 {config_name} 至本地雲端目錄。"
                WSysLog("1", "ConfigFileToCould", msg)
    except (IOError, shutil.Error) as e:
        msg = f"拷貝 {config_name} 至本地雲端目錄發生錯誤。錯誤原因： {e}。"
        WSysLog("3", "ConfigFileToCould", msg)

# 判斷當天是否為工作日，寫入System.json檔案
def WorkDay():
    cal = Taiwan()
    result = cal.is_working_day(Config.SystemTime)
    msg = WrtieWorkDay(result)
    WSysLog("1", "WorkDay", msg)

if __name__ == "__main__":
    # CheckBAFolderFiles()
    CheckFileInfo()
    # aa = ["BA01","BA02","BA03","BA04"]
    # MergeDealerInfo(aa)
    # MergeDealerInfoWorkSheet(["BA01"])
    # RenewDealerJson()
    # ConfigFile()
    # MergeMasterFile(["BA03"])
    # MergeKalist(["BA03"])
    # WorkDay()
    # print(Config.WorkDay)