# -*- coding: utf-8 -*-

"""
檔案說明：撰寫繳交紀錄表
Writer：Qian
"""

"""
繳交紀錄表
    write_data{
        UploadData{
            "經銷商ID"
            "檔案類型"
            "繳交狀態"
            "檔案名稱"
            "應繳時間"
            "繳交時間"
            "檔案內容總筆數"
        }
    }
    write_data{
        CheckData{
            "ID"
            "檢查狀態"
            "表頭檢查結果"
            "內容檢查結果"
            "內容錯誤筆數"
        }
    }
    write_data{
        ChangeData{
            "ID"
            "轉換狀態"
            "轉換後檔案名稱"
            "轉換錯誤筆數"
            "轉換後總筆數"
        }
    }
每月總結紀錄表
    write_data{
        "經銷商ID",
        "檔案類型",
        "當月繳交次數"
        "當月繳交筆數"
        "當月繳交錯誤次數"
        "當月繳交錯誤筆數"
        "當月轉換次數"
        "當月轉換筆數"
        "當月轉換錯誤次數"
        "當月轉換錯誤筆數"
    }

待補繳紀錄表
    新增資料
    write_data{
        "經銷商ID"  key
        "檔案類型"  key
        "缺繳(待補繳)檔案名稱" key
        "檔案狀態"
        "應繳時間"
        "檔案檢查結果"
    }
    更新資料
    write_data{
        "經銷商ID"  key
        "檔案類型"  key
        "缺繳(待補繳)檔案名稱" key
        "檔案狀態"
        "補繳時間"
        "補繳檢查結果"
    }
"""

# 標準庫
import os, re

# 第三方庫
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

# 自定義函數
from Mail import SendMail
from Config import AppConfig
from Log import WSysLog, WRecLog
from SystemConfig import SubRecordJson

Config = AppConfig()

# 報表共用
# 產生對應的 Excel Column 名稱
def get_excel_colmun_name(file_header):
    return [chr(i % 26+ 65) for i in range(len(file_header))]

# 設定表格樣式
def excel_style(ws, column_width, fixed_columns):
    for i, width in enumerate(column_width):
        ws.column_dimensions[fixed_columns[i]].width = width
    for col in fixed_columns:
        for row in range(1, ws.max_row + 1):
            ws[f"{col}{row}"].alignment = Config.ExcelStyle

# 建立表頭
def make_header(wb, file_sheet_name, file_header, file_path, column_width):
    fixed_columns = get_excel_colmun_name(file_header)
    ws = wb.create_sheet(title = file_sheet_name)
    for col, data in zip(fixed_columns, file_header):
        ws[f"{col}1"] = data
    excel_style(ws, column_width, fixed_columns)
    wb.save(file_path)

# 每日、每月總結紀錄表用
# 寫入經銷商資訊
def write_dealer_info(file_path, file_name, file_sheet_name, file_header):
    wb = load_workbook(file_path)
    ws = wb[file_sheet_name]
    dealer_data = {}
    # 取得資料內容
    for i in range(len(Config.DealerList)):
        index  = i + 1
        dealer_id = Config.DealerConfig[f"Dealer{index}"]["DealerID"]
        dealer_name = Config.DealerConfig[f"Dealer{index}"]["DealerName"]
        sale_cycle = Config.DealerConfig[f"Dealer{index}"]["SaleFile"]["PaymentCycle"]
        inventory_cycle = Config.DealerConfig[f"Dealer{index}"]["InventoryFile"]["PaymentCycle"]
        dealer_data[i] = {"經銷商ID":dealer_id, "經銷商名稱":dealer_name, "檔案類型":"Sale",\
                          "檔案繳交週期":sale_cycle,"檔案類型1":"Inventory","檔案繳交週期1":inventory_cycle}
    
    # 寫入經銷商資訊
    for i in range(2,len(Config.DealerList)*2+2,2):
        index = int((i / 2) -1)
        for col_name, input_data in dealer_data[index].items():
            row = i
            if col_name == "檔案類型1":
                col_name = "檔案類型"
                row += 1
            if col_name == "檔案繳交週期1":
                col_name = "檔案繳交週期"
                row += 1
            col = search_column_name(file_header, col_name)
            ws[f"{col}{row}"] = input_data
            ws[f"{col}{row}"].alignment = Config.ExcelStyle
        ws.merge_cells(f"A{i}:A{i +1}")
        ws.merge_cells(f"B{i}:B{i +1}")
    wb.save(file_path)
    msg = f"已將經銷商資訊寫入 {file_name} 中。"
    WRecLog("1", "WriteDealerInfo", "All Dealer", file_name, msg)

# 報表共用
# 製作繳交紀錄表 RawData
def check_templates(mode):
    if mode == "SubRawData":
        file_path = Config.SubRawDataPath
        file_name = Config.SubRawDataFileName
        file_sheet_name = Config.SubRawDataSheetName
        file_header = Config.SubRawDataHeader
        column_width = Config.SubRawDataColumnWidth
    elif mode == "Daily":
        file_path = Config.DailyReportPath
        file_name = Config.DailyReportFileName
        file_sheet_name = Config.DailyReportSheetName
        file_header = Config.DailyReportHeader
        column_width = Config.DailyReportColumnWidth
    elif mode == "Monthly":
        file_path = Config.MonthlyReportPath
        file_name = Config.MonthlyReportFileName
        file_sheet_name = Config.MonthlyReportSheetName
        file_header = Config.MonthlyReportHeader
        column_width = Config.MonthlyReportColumnWidth
    elif mode == "NotSub":
        file_path = Config.NotSubPath
        file_name = Config.NotSubFileName
        file_sheet_name = Config.NotSubSheetName
        file_header = Config.NotSubHeader
        column_width = Config.NotSubColumnWidth
    else:
        msg = "輸入的mode未在規範中。"
        WSysLog("3", "CheckTemplates", msg)
        return False
    
    if os.path.exists(Config.ReportFolderPath):
        try:
            wb = load_workbook(file_path)
            if file_sheet_name in wb.sheetnames:
                msg = f"已確認 {file_sheet_name} 工作表，存在於 {file_name} 檔案中。"
                WSysLog("1", "CheckTemplates", msg)
            else:
                make_header(wb, file_sheet_name, file_header, file_path, column_width)
                if (mode == "Daily") or (mode == "Monthly"):
                    write_dealer_info(file_path, file_name, file_sheet_name, file_header)
                msg = f"成功在檔案 {file_name} 中，建立工作表 {file_sheet_name}。"
                WSysLog("1", "CheckTemplates", msg)
            return True
        except FileNotFoundError:
            wb = Workbook()
            wb.remove(wb.active)
            make_header(wb, file_sheet_name, file_header, file_path, column_width)
            if (mode == "Daily") or (mode == "Monthly"):
                write_dealer_info(file_path, file_name, file_sheet_name, file_header)
            msg = f"成功建立 {file_name} 檔案。"
            WSysLog("1", "CheckTemplates", msg)
            return True
    else:
        msg = f"{Config.ReportFolderPath} 目錄不存在。"
        WSysLog("3", "CheckTemplates", msg)
        return False

# 根據表頭搜尋出 Excel 的 Column 名稱
def search_column_name(file_header, col_name):
    for index in range(len(file_header)):
        if file_header[index] == col_name:
            break
    return chr(index % 26 + 65)

# SubRawData
# 寫入繳交資料
def write_upload_data(ws, data, file_header):
    file_upload_time = data[file_header[8]]
    file_name = data[file_header[6]]
    row = None
    col = search_column_name(file_header, file_header[8])
    for cell in ws[col]:
        if cell.value == file_upload_time:
            col = search_column_name(file_header, file_header[6])
            if ws[f"{col}{cell.row}"].value == file_name:
                row = cell.row
                data_id = ws[f"A{cell.row}"].value
                break
    if not row:
        row = ws.max_row +1
        data_id = row -1
    for i in range(len(Config.DealerList)):
        if Config.DealerList[i] == data["經銷商ID"]:
            index = i + 1
            break
    # 搜尋檔案繳交週期
    sale_cycle = Config.DealerConfig [f"Dealer{index}"]["SaleFile"]["PaymentCycle"]
    inventory_cycle = Config.DealerConfig[f"Dealer{index}"]["InventoryFile"]["PaymentCycle"]
    sub_cycle = sale_cycle if data["檔案類型"] == "Sale" else inventory_cycle

    # 寫入ID、經銷商名稱、檔案繳交週期
    data["ID"] = data_id    
    data["經銷商名稱"] = Config.DealerConfig[f"Dealer{index}"]["DealerName"]
    data["檔案繳交週期"] = sub_cycle

    # 將字典資料寫入檔案
    for col_name, input_data in data.items():
        col = search_column_name(file_header, col_name)
        ws[f"{col}{row}"] = input_data
        ws[f"{col}{row}"].alignment = Config.ExcelStyle
    
    return data_id

# 寫入檢查資料
def write_check_or_change_data(ws, data, file_header):
    data_id = data["ID"]
    row = int(data_id) + 1
    data.pop("ID")
    for col_name, input_data in data.items():
        col = search_column_name(file_header, col_name)
        ws[f"{col}{row}"] = input_data
        ws[f"{col}{row}"].alignment = Config.ExcelStyle
    return data_id

# 撰寫繳交紀錄表
def WriteSubRawData(write_data):
    file_path = Config.SubRawDataPath
    file_name = Config.SubRawDataFileName
    file_sheet_name = Config.SubRawDataSheetName
    file_header = Config.SubRawDataHeader
    result = check_templates("SubRawData")
    if result:
        wb = load_workbook(file_path)
        ws = wb[file_sheet_name]
        if write_data == "Read":
            return True, ws.max_row
        # 寫入UploadData
        if "UploadData" in write_data:
            data = write_data["UploadData"]
            data_id = write_upload_data(ws, data, file_header)
            wb.save(file_path)
            msg = f"{file_sheet_name} 工作表中，新增資料，ID： {data_id}。"
            WRecLog("1", "WrightSubRawData", "All Dealer", file_name, msg)
            return True, data_id
        # 寫入CheckData
        elif "CheckData" in write_data:
            data = write_data["CheckData"]
            data_id = write_check_or_change_data(ws, data, file_header)
            wb.save(file_path)
            msg = f"{file_sheet_name} 工作表中，更新 ID：{data_id} 資料：{data}。"
            WRecLog("1", "WrightSubRawData", "All Dealer", file_name, msg)
            return True, data_id
        # 寫入ChangeData
        elif "ChangeData" in write_data:
            data = write_data["ChangeData"]
            data_id = write_check_or_change_data(ws, data, file_header)
            wb.save(file_path)
            msg = f"{file_sheet_name} 工作表中，更新 ID：{data_id} 資料：{data}。"
            WRecLog("1", "WrightSubRawData", "All Dealer", file_name, msg)
            return True, data_id
        else:
            msg = f"寫入的資料不符合規範，{write_data}。"
            WRecLog("2", "WrightSubRawData", "All Dealer", file_name, msg)
            return False, None
    else:
        msg = "工作表新增資料失敗。"
        WRecLog("2", "WriteSubRawData", "All Dealer", file_name, msg)
        return result, None

# 撰寫每日總結紀錄表
def WriteDailyReoprt(write_data):
    file_path = Config.DailyReportPath
    file_name = Config.DailyReportFileName
    file_sheet_name = Config.DailyReportSheetName
    column_name = f"{Config.Month}月{Config.Day}日"
    column_width = Config.DailyReportNewDataWidth
    result = check_templates("Daily")
    if result:
        wb = load_workbook(file_path)
        ws = wb[file_sheet_name]
        column_index = None
        for col in range(1, ws.max_column + 1):
            if ws.cell(row = 1, column = col).value == column_name:
                column_index = col
                break

        # 搜索表頭對應欄位與搜尋、建立新表頭
        if column_index is None:
            column_index = ws.max_column + 1
            ws.cell(row = 1, column = column_index, value = column_name)
            column_str = get_column_letter(column_index)
            ws.column_dimensions[column_str].width = column_width
            ws[f"{column_str}1"].alignment = Config.ExcelStyle
            msg = f"{file_sheet_name} 工作表，新增資料 {column_name} ： {write_data}。"
            WRecLog("1", "WriteDailyReoprt", "All Dealer", file_name, msg)
        else:
            column_str = get_column_letter(column_index)
            msg = f"{file_sheet_name} 工作表，{column_name} 更新資料  ： {write_data}。"
            WRecLog("1", "WriteDailyReoprt", "All Dealer", file_name, msg)
        
        # 寫入當天資料
        for row, input_data in enumerate(write_data, start = 2):
            ws.cell(row = row, column = column_index, value = input_data)
            ws[f"{column_str}{row}"].alignment = Config.ExcelStyle
        
        # 取出檔案總筆數
        update_num = [re.split(r"[:;：；/]", input_data)[2] for input_data in write_data]
        
        # 寫入當天更新筆數欄位
        for row, input_data in enumerate(update_num, start = 2):
            ws.cell(row = row, column = 5, value = input_data)
            ws[f"E{row}"].alignment = Config.ExcelStyle

        wb.save(file_path)
        return True
    else:
        msg = "工作表新增/更新資料失敗。"
        WRecLog("2", "WriteDailyReoprt", "All Dealer", file_name, msg)
        return result
    
# 撰寫每月總結紀錄表
def WriteMonthlyReoprt(write_data):
    file_path = Config.MonthlyReportPath
    file_name = Config.MonthlyReportFileName
    file_sheet_name = Config.MonthlyReportSheetName
    file_header = Config.MonthlyReportHeader
    result = check_templates("Monthly")
    if result:
        wb = load_workbook(file_path)
        ws = wb[file_sheet_name]
        
        dealer_id = write_data["經銷商ID"]
        data_type = write_data["檔案類型"]
        # 抓取row
        for i in range(len(Config.DealerList)):
            if dealer_id == Config.DealerList[i]:
                row = (i + 1) * 2
                break
        if data_type == "Inventory":
            row += 1
        write_data.pop("經銷商ID")
        write_data.pop("檔案類型")
        for col_name, input_data in write_data.items():
            col = search_column_name(file_header, col_name)
            old_data = ws[f"{col}{row}"].value
            if old_data is None:
                old_data = 0
            ws[f"{col}{row}"] = old_data + input_data
            ws[f"{col}{row}"].alignment = Config.ExcelStyle
            write_data[col_name] = old_data + input_data
        msg = f"{file_sheet_name} 工作表中，經銷商ID：{dealer_id}，檔案類型：{data_type}，更新資料：{write_data}"
        WRecLog("1", "WriteMonthlyReoprt", "All Dealer", file_name, msg)
        wb.save(file_path)
        return True
    else:
        msg = "工作表新增/更新資料失敗。"
        WRecLog("2", "WriteMonthlyReoprt", "All Dealer", file_name, msg)
        return result

# 撰寫待補繳紀錄表
def WriteNotSubmission(write_data):
    file_path = Config.NotSubPath
    file_name = Config.NotSubFileName
    file_sheet_name = Config.NotSubSheetName
    file_header = Config.NotSubHeader
    result = check_templates("NotSub")
    new_data_flag = False
    if result:
        # 讀取未繳交紀錄表中的內容
        # 日繳
        if write_data == "ReadDaily":
            df = pd.read_excel(file_path, sheet_name = file_sheet_name, dtype = str, index_col = "ID")
            df = df[df[Config.NotSubHeader[4]] == "D"]
            return df
        # 月繳
        if write_data == "ReadMonthly":
            df = pd.read_excel(file_path, sheet_name = file_sheet_name, index_col = "ID")
            df = df[df[Config.NotSubHeader[4]] == "M"]
            return df
        
        wb = load_workbook(file_path)
        ws = wb[file_sheet_name]
        dealer_id = write_data["經銷商ID"]
        data_type = write_data["檔案類型"]
        for i in range(len(Config.DealerList)):
            if dealer_id == Config.DealerList[i]:
                index = i + 1
                break
        dealer_name = Config.DealerConfig[f"Dealer{index}"]["DealerName"]
        sub_cycle = Config.DealerConfig[f"Dealer{index}"][f"{data_type}File"]["PaymentCycle"]
        column = search_column_name(file_header, "缺繳(待補繳)檔案名稱")
        not_sub_file_name = [cell.value for cell in ws[column]]
        row = None

        # 更新資料
        for i in range(len(not_sub_file_name)):
            if write_data["缺繳(待補繳)檔案名稱"] == not_sub_file_name[i]:
                write_data.pop("經銷商ID")
                write_data.pop("檔案類型")
                write_data.pop("缺繳(待補繳)檔案名稱")
                row = i + 1
                break

        # 新增資料
        if row is None:
            write_data["ID"] = ws.max_row
            write_data["經銷商名稱"] = dealer_name
            write_data["檔案繳交週期"] = sub_cycle
            new_data_flag = True
            row = ws.max_row + 1

        # 寫入資料
        for col_name, input_data in write_data.items():
            col = search_column_name(file_header, col_name)
            ws[f"{col}{row}"] = input_data
            ws[f"{col}{row}"].alignment = Config.ExcelStyle

        wb.save(file_path)
        if new_data_flag:
            msg = f"{file_sheet_name} 工作表中，新增資料，ID：{row -1}，{write_data}。"
        else:
            msg = f"{file_sheet_name} 工作表中，ID：{row - 1}，更新資料{write_data}。"
        WRecLog("1", "WriteNotSubmission", "All Dealer", file_name, msg)
        return True
    else:
        msg = "工作表新增/更新資料失敗。"
        WRecLog("2", "WriteNotSubmission", "All Dealer", file_name, msg)
        return result

# 抓取繳交紀錄表資料，轉換寫入每日總結紀錄表
def statistics_daily_data(df, file_header):
    daily_data = []
    for i in range(len(Config.DealerList)):
        for j in range(2):
            num_data = []
            data_type = "Sale" if j == 0 else "Inventory"
            sort_data =  df[(df[file_header[1]] == Config.DealerList[i]) & (df[file_header[3]] == data_type)]
            sub_data = sort_data[sort_data[file_header[5]] == "有繳交"]
            resub_data = sort_data[sort_data[file_header[5]] == "補繳交"]
            for k in range(2):
                data = sub_data if k == 0 else resub_data
                total_content_num = 0
                for num in data[file_header[9]].dropna():
                    total_content_num += int(num)
                
                # num_data[0]、[2]
                num_data.append(len(data))
                
                # num_data[1]、[3]
                num_data.append(total_content_num)
            
            # num_data[4]
            num_data.append(len(sort_data[sort_data[file_header[10]] == "OK"]))
            
            # num_data[5]
            num_data.append(len(sort_data[sort_data[file_header[10]] == "NO"]))
            total_content_num = 0
            for num in sort_data[file_header[13]].dropna():
                total_content_num += int(num)
            
            # num_data[6]
            num_data.append(total_content_num)
            
            # num_data[7]
            num_data.append(len(sort_data[file_header[14]].dropna()))
            
            total_content_num = 0
            for num in sort_data[file_header[16]].dropna():
                total_content_num += int(num)
            
            ## num_data[8]
            num_data.append(total_content_num)
            
            # 繳交：檔案數/資料總比數；補繳：檔案數/資料總比數；檢查：檢查正確檔案數/檢查錯誤檔案數/內容錯誤總筆數；轉換:轉換總檔案數/轉換錯誤總筆數
            write_data = f"繳交：{num_data[0]}/{num_data[1]}；補繳：{num_data[2]}/{num_data[3]}；檢查：{num_data[4]}/{num_data[5]}/{num_data[6]}；轉換：{num_data[7]}/{num_data[8]}"
            daily_data.append(write_data)
    return daily_data

# 抓取繳交紀錄表資料，轉換寫入每月總結紀錄表
def statistics_monthly_data(df,sub_file_header, monthly_file_header):
    for i in range(len(Config.DealerList)):
        for j in range(2):
            content_num, content_error_num, change_num, change_error_num = 0, 0, 0, 0
            data_type = "Sale" if j == 0 else "Inventory"
            sort_data =  df[(df[sub_file_header[1]] == Config.DealerList[i]) & (df[sub_file_header[3]] == data_type)]
            for num in sort_data[sub_file_header[9]].dropna():
                content_num += int(num)
            for num in sort_data[sub_file_header[13]].dropna():
                content_error_num += int(num)
            change_data = sort_data[sort_data[sub_file_header[10]] == "OK"]
            for num in change_data[sub_file_header[9]].dropna():
                change_num += int(num)
            change_error_data = sort_data[(sort_data[sub_file_header[10]] == "OK") & (sort_data[sub_file_header[14]] == "NO")]
            if not change_error_data[sub_file_header[16]].dropna().empty:
                for num in change_error_data[sub_file_header[16]].dropna():
                    change_error_num += int(num)
            monthly_data = {
                monthly_file_header[0]:Config.DealerList[i],
                monthly_file_header[2]:data_type,
                monthly_file_header[4]:len(sort_data[sub_file_header[5]].dropna()),
                monthly_file_header[5]:content_num,
                monthly_file_header[6]:len(sort_data[sort_data[sub_file_header[10]] == "NO"]),
                monthly_file_header[7]:content_error_num,
                monthly_file_header[8]:len(sort_data[sub_file_header[14]].dropna()),
                monthly_file_header[9]:change_num,
                monthly_file_header[10]:len(sort_data[sort_data[sub_file_header[14]] == "NO"]),
                monthly_file_header[11]:change_error_num
            }
            WriteMonthlyReoprt(monthly_data)

# 取得轉換報告之信件參數
def WriteChangeReportMail(df, file_header):
    mail_data_num = [0] * 6
    for i in range(len(Config.DealerList)):
        for j in range(2):
            num_data = []
            data_type = "Sale" if j == 0 else "Inventory"
            sort_data =  df[(df[file_header[1]] == Config.DealerList[i]) & (df[file_header[3]] == data_type)]
            sub_data = sort_data[sort_data[file_header[5]] == "有繳交"]
            resub_data = sort_data[sort_data[file_header[5]] == "補繳交"]
            for k in range(2):
                data = sub_data if k == 0 else resub_data
                total_content_num = 0
                for num in data[file_header[9]].dropna():
                    total_content_num += int(num)
                
                # num_data[0]、[2]
                num_data.append(len(data))
                
                # num_data[1]、[3]
                num_data.append(total_content_num)

            # num_data[4]
            total_content_num = 0
            for num in sort_data[file_header[13]].dropna():
                total_content_num += int(num)
            num_data.append(total_content_num)

            ## num_data[5]
            total_content_num = 0
            for num in sort_data[file_header[16]].dropna():
                total_content_num += int(num)
            num_data.append(total_content_num)
            mail_data_num[0] = mail_data_num[0] + num_data[0]
            mail_data_num[1] = mail_data_num[1] + num_data[1]
            mail_data_num[2] = mail_data_num[2] + num_data[2]
            mail_data_num[3] = mail_data_num[3] + num_data[3]
            mail_data_num[4] = mail_data_num[4] + num_data[4]
            mail_data_num[5] = mail_data_num[5] + num_data[5]
    mail_data_num[0] = mail_data_num[0] + mail_data_num[2]
    mail_data_num[1] = mail_data_num[1] + mail_data_num[3]
    del mail_data_num[3]
    del mail_data_num[2]
    # 繳交檔案總數；繳交檔案總比數；內容錯誤總筆數；轉換錯誤筆數

    report_name = [Config.DailyReportFileName, Config.MonthlyReportFileName]
    mail_data = {"FileNum" : mail_data_num[0], "DataNum" : mail_data_num[1], "CheckErrorNum" : mail_data_num[2],
            "ChangeErrorNum" : mail_data_num[3], "ReportName": "、".join(report_name)}
    report_paths = [Config.DailyReportPath, Config.MonthlyReportPath]
    send_info = {"Mode" : "ChangeReport", "DealerID" : None, "MailData" : mail_data, "FilesPath" : report_paths}
    SendMail(send_info)

# 讀取 SubRawData 資料
def Statistics():
    # 取得資料範圍    
    start_index = SubRecordJson("ReadSubStartIndex", None)
    _, end_index = WriteSubRawData("Read")
    end_index = end_index - 1
    # 取得資料
    file_path = Config.SubRawDataPath
    file_sheet_name = Config.SubRawDataSheetName
    sub_file_header = Config.SubRawDataHeader
    monthly_file_header = Config.MonthlyReportHeader
    index_key = "ID"
    df = pd.read_excel(file_path, sheet_name = file_sheet_name, dtype = str, index_col = index_key)
    df = df[start_index:end_index]
    # Daily
    daily_data =  statistics_daily_data(df, sub_file_header)
    WriteDailyReoprt(daily_data)
    msg = "已將繳交紀錄資訊寫入至每日總結紀錄表。"
    WSysLog("1", "Statistics", msg)
    # Monthly
    statistics_monthly_data(df,sub_file_header, monthly_file_header)
    msg = "已將繳交紀錄資訊寫入至每月總結紀錄表。"
    WSysLog("1", "Statistics", msg)
    WriteChangeReportMail(df, sub_file_header)
    # 更新json中的rawdata起始值
    msg = SubRecordJson("WriteSubStartIndex", end_index)
    WSysLog("1", "Statistics", msg)
    
if __name__ == "__main__":
    # data0 = {"UploadData":{
    #     "經銷商ID":"111",
    #     "檔案類型":"Sale",
    #     "繳交狀態":"有繳交",
    #     "檔案名稱":"0.xlsx",
    #     "應繳時間":"2024-07-18 22:00",
    #     "繳交時間":"2024-07-18 15:58",
    #     "檔案內容總筆數":"500"
    # }}
    # data1 = {"CheckData":{
    #     "ID":"1",
    #     "檢查狀態":"OK",
    #     "表頭檢查結果":"good",
    #     "內容檢查結果":"good",
    #     "內容錯誤筆數":"0",
    # }}
    # data2 = {"ChangeData":{
    #     "ID":"1",
    #     "轉換狀態":"OK",
    #     "轉換後檔案名稱":"0.xlsx",
    #     "轉換錯誤筆數":"0",
    #     "轉換後總筆數":"500"
    # }}
    # WriteSubRawData(data2)
    # data0 = ["繳交：1/10；補繳：0/0；檢查：True/2；轉換：1/52"]
    # WriteDailyReoprt(data0)
    # data0 = {"經銷商ID":"111","檔案類型":"Inventory","當月繳交次數":1,"當月繳交筆數":1,"當月繳交錯誤次數":10,"當月繳交錯誤筆數":10,"當月轉換次數":5,"當月轉換筆數":6,"當月轉換錯誤次數":4,"當月轉換錯誤筆數":4}
    # WriteMonthlyReoprt(data0)
    # data0 = {
    #     "經銷商ID":"111",
    #     "檔案類型":"Sale",
    #     "缺繳(待補繳)檔案名稱":"0.xlsx",
    #     "檔案狀態":"00",
    #     "應繳時間":"eee",
    #     "檔案檢查結果":"aaa",
    #     "補繳時間":"hhh",
    #     "補繳檢查結果":"333"
    # }
    # WriteNotSubmission(data0)
    Statistics()
    # print(Config.DailyReportFileName)